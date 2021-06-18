# -*- coding=UTF-8 -*-
# pyright: strict
""".  """

from __future__ import (absolute_import, division, print_function,
                        unicode_literals)

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from typing import Dict, Text, Tuple, Any, Optional, Iterator
import cast_unknown as cast
import logging
import six

LOGGER = logging.getLogger(__name__)


def _column_by_alias(header, header_alias):
    # type: (Any, Dict[Text, Tuple[Text, ...]]) -> Text
    for k, v in six.iteritems(header_alias):
        if header in v:
            return k
    return cast.text(header)

def _columns(header_row, header_alias):
    # type: (Tuple[Cell, ...], Dict[Text, Tuple[Text, ...]]) -> Tuple[Text, ...]
    """Find sheet field name for each column.  """

    return tuple(_column_by_alias(i.value, header_alias) for i in header_row)


def iter_rows_as_dict(rows, header_alias=None):
    # type: (Iterator[Tuple[Cell, ...]], Optional[Dict[Text, Tuple[Text, ...]]]) -> Iterator[Dict[Text, Any]]
    columns = ()
    for row in rows:
        # ignore empty row
        if all(cell.value is None for cell in row):
            continue
        if not len(columns):
            columns = _columns(row, header_alias or {})
        else:
            yield dict((columns[col_index], cell.value) for col_index,cell in enumerate(row))

def iter_workbook_as_dict(workbook, header_alias=None):
    # type: (Workbook, Optional[Dict[Text, Tuple[Text, ...]]]) -> Iterator[Dict[Text, Any]]
    for sheet in workbook.worksheets:
        for i in iter_rows_as_dict(sheet.iter_rows(), header_alias):
            yield i

