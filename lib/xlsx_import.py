# -*- coding=UTF-8 -*-
"""Import xlsx sheet to cgtw database then set status.  """

from __future__ import (absolute_import, division, print_function,
                        unicode_literals)

import logging
from collections import namedtuple

import openpyxl
import six
import win_unicode_console
from Qt.QtWidgets import QFileDialog

import cgtwq
from wlf.console import pause
from wlf.progress import CancelledError, progress
from wlf.uitools import application

__version__ = '1.4.2'
LOGGER = logging.getLogger(__name__)

HEAD_ALIAS = {
    'pipeline': ['流程', '流程名', '制作流程'],
    'shot': ['镜头', '镜头号', '镜头名'],
    'phase': ['阶段', '返修阶段'],
    'status': ['状态', '返修状态'],
    'note': ['备注', '说明', '内容', '描述']
}
METHOD_ALIAS = {
    'retake': ['返修'],
    'approve': ['通过'],
}
FIELD_ALIAS = {
    'leader_status': ['组长', '组长状态'],
    'director_status': ['导演', '导演状态'],
    'client_status': ['客户', '客户状态'],
}


class RowData(namedtuple('RowData', ('pipeline', 'shot',
                                     'phase', 'status', 'note',
                                     'index'))):
    """Data in a row.  """
    head_fields = ('pipeline', 'shot',
                   'phase', 'status', 'note')


def ask_filename():
    """Ask user for filename.  """

    filename, _ = QFileDialog.getOpenFileName(
        None, caption='选择要读取的文件…', filter='*.xlsx')
    return filename


def get_data(filename):
    """Get data from file.  """
    ret = []
    for i in openpyxl.load_workbook(filename).worksheets:
        try:
            ret.extend(parse_sheet(i))
        except ParseException as ex:
            if isinstance(ex, EmptySheet):
                continue
            LOGGER.error('工作表分析失败: %s', i.title)
    return ret


def parse_sheet(sheet):
    """Parse datasheet"""
    assert isinstance(sheet, openpyxl.worksheet.Worksheet)

    first_row = find_first_row(sheet)
    data_start_row = first_row[0].row + 1
    data_columns = find_data_columns(first_row)

    ret = []

    for row in progress(
            sheet.iter_rows(min_row=data_start_row),
            total=sheet.max_row - data_start_row + 1):
        data = parse_row(row, data_columns)
        if data:
            ret.append(data)

    return ret


class ParseException(Exception):
    """Base class for parse exceptions.  """


class EmptySheet(ParseException):
    """Indicate the sheet is empty.  """


class InvalidSheet(ParseException):
    """Indicate the sheet is invalid.  """


class ImportException(Exception):
    """Base class for import exceptions.  """


class NoEntry(ImportException):
    """Indicate no entry in database.  """


def find_first_row(sheet):
    """Find first avaliable row in sheet.  """

    assert isinstance(sheet, openpyxl.worksheet.Worksheet)
    for row in sheet.rows:
        # Ignore empty row.
        if all(cell.value is None for cell in row):
            continue
        return row
    LOGGER.warning('工作表是空的: %s', sheet.title)
    raise EmptySheet(sheet.title)


def find_data_columns(row):
    """Find sheet column indexes for requried data.  """

    return RowData(index=row[0].row,
                   **{i: _find_column(row, i) for i in RowData.head_fields})


def _find_column(row, key):
    for cell in row:
        assert isinstance(cell, openpyxl.cell.Cell)
        if _convert_from_alias(cell.value, HEAD_ALIAS) == key:
            return cell.col_idx
    LOGGER.error('在表头中找不到字段: %s, 支持的别称为: %s',
                 key, ', '.join(HEAD_ALIAS[key]))
    raise InvalidSheet(key)


def parse_row(row, data_columns):
    """Get required data from row.  """

    assert isinstance(data_columns, RowData)
    data = {}
    for field in RowData.head_fields:
        column_index = getattr(data_columns, field) - 1
        data[field] = row[column_index].value
    return RowData(index=row[0].row, **data)


def import_data(data, database, module, module_type):
    """Import data to cgtw database.  """

    module = cgtwq.Database(database).module(module, module_type)
    errors = []
    for i in data:
        assert isinstance(i, RowData)
        try:
            select = module.filter(
                (cgtwq.Field('shot.shot') == i.shot)
                & (cgtwq.Field('pipeline') == i.pipeline))
        except ValueError:
            LOGGER.error('找不到对应条目: 行=%s, 数据库=%s, 镜头号=%s, 流程=%s',
                         i.index, database, i.shot, i.pipeline, )
            errors.append(NoEntry(i))

        try:
            _apply_on_selection(select, i)
        except ValueError:
            continue
    if errors:
        raise ImportError(errors)


def _apply_on_selection(select, data):
    def _check_alias(value, alias, name, label):
        if value not in alias:
            LOGGER.error('不能识别表格中给出的任务%s: 行=%s, 值=%s, 支持的值有: %s',
                         label, data.index, value,
                         ', '.join(i for j in [alias.keys()] + alias.values() for i in j))
            raise ValueError('Can not recognize {}.'.format(name), value)

    assert isinstance(data, RowData)
    assert isinstance(select, cgtwq.Selection)
    method = _convert_from_alias(data.status, METHOD_ALIAS)
    _check_alias(method, METHOD_ALIAS, 'status', '状态')
    field = _convert_from_alias(data.phase, FIELD_ALIAS)
    _check_alias(field, FIELD_ALIAS, 'phase', '阶段')
    if field not in FIELD_ALIAS:
        LOGGER.error('不能识别表格中给出的任务阶段: 行=%s, 值=%s, 支持的值有: %s', data.index, field,
                     ', '.join(i for j in [FIELD_ALIAS.keys()] + FIELD_ALIAS.values() for i in j))
        raise ValueError('Can not recognize field.', field)

    entries = select.to_entries()
    if len(entries) != 1:
        LOGGER.warning('发现%s个名为%s的%s任务, 都将被修改',
                       len(entries), data.shot, data.pipeline)

    for entry in select.to_entries():
        assert isinstance(entry, cgtwq.Entry)

        if data.note:
            message = cgtwq.Message(_convert_note(data.note))
            _message = message.dumps()
            if entry.history.get(
                    (cgtwq.Field('status') == method.capitalize())
                    & (cgtwq.Field('text') == _message)):
                LOGGER.info('此条已导入, 跳过: 行=%s', data.index)
                continue

        try:
            if method == 'retake':
                entry.flow.retake(field, message=message)
            elif method == 'approve':
                entry.flow.approve(field, message=message)
            LOGGER.info('成功导入: 行=%s, 镜头=%s, 流程=%s, 阶段=%s, 状态=%s, 备注=%s',
                        data.index, data.shot, data.pipeline, data.phase, data.status, data.note)
        except cgtwq.PermissionError:
            LOGGER.error('当前用户无权限: 行=%s, 阶段=%s', data.index, data.phase)


def _convert_note(value):
    if not value:
        return value
    _value = six.text_type(value)
    _value = _value.replace('\n', '<br>')
    return _value


def _convert_from_alias(value, alias_dict):
    if not value:
        return value
    _value = six.text_type(value)
    _value = _value.lower()
    try:
        return next(i for i in alias_dict if i == _value or _value in alias_dict[i])
    except StopIteration:
        return value


def main():
    win_unicode_console.enable()
    logging.basicConfig(
        level='INFO', format='%(levelname)-7s:%(name)s: %(message)s')

    print('{:-^50}'.format('导入XLSX返修表 v{}'.format(__version__)))
    dummy = application()
    client = cgtwq.DesktopClient()
    client.connect()
    plugin_data = client.plugin.data()

    filename = ask_filename()
    if not filename:
        return

    try:
        data = get_data(filename)
        if not data:
            LOGGER.error('没能发现任何可用数据')
            raise ParseException
        import_data(data,
                    plugin_data.database,
                    plugin_data.module,
                    plugin_data.module_type)
        pause()
    except CancelledError:
        LOGGER.info('用户取消')
    except ParseException:
        print('\n分析表格失败, 请检查上方日志\n')
        pause(0)
        return
    except ImportException:
        print('\n导入数据失败, 请检查上方日志\n')
        pause(0)
        return
    except:
        import traceback
        traceback.print_exc()
        pause(0)
        raise


if __name__ == '__main__':
    main()
